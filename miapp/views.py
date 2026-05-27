from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.models import User
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from .forms import RegistroForm, LoginForm, PedidoForm
from .models import Cliente, Pedido
from .repositories import ClienteRepository, PedidoRepository, UsuarioRepository
import requests
import io


# ─────────────────────────────────────────
# PÁGINAS GENERALES
# ─────────────────────────────────────────

def inicio_sin_sesion(request):
    """
    Página de inicio pública. Carga 4 libros de la API para
    mostrar en la sección Colecciones Destacadas.
    """
    libros_destacados = []
    try:
        headers  = {'User-Agent': 'Mozilla/5.0'}
        url      = "https://openlibrary.org/search.json?q=don+quijote+garcia+marquez+neruda+borges&limit=4&lang=spa"
        response = requests.get(url, timeout=6, headers=headers)
        if response.status_code == 200:
            data = response.json()
            for doc in data.get('docs', [])[:4]:
                cover_id = doc.get('cover_i')
                libros_destacados.append({
                    'titulo': doc.get('title', 'Sin título'),
                    'autor':  (doc.get('author_name') or ['Desconocido'])[0],
                    'imagen': f"https://covers.openlibrary.org/b/id/{cover_id}-M.jpg" if cover_id else '',
                })
    except Exception:
        pass

    # Fallback si la API no responde
    if not libros_destacados:
        libros_destacados = [
            {'titulo': 'Cien Años de Soledad',  'autor': 'Gabriel García Márquez', 'imagen': 'https://covers.openlibrary.org/b/id/8224600-M.jpg'},
            {'titulo': 'La Metamorfosis',        'autor': 'Franz Kafka',            'imagen': 'https://covers.openlibrary.org/b/id/8091024-M.jpg'},
            {'titulo': 'Veinte Poemas de Amor',  'autor': 'Pablo Neruda',           'imagen': 'https://covers.openlibrary.org/b/id/8226198-M.jpg'},
            {'titulo': 'Don Quijote',            'autor': 'Miguel de Cervantes',    'imagen': 'https://covers.openlibrary.org/b/id/8739161-M.jpg'},
        ]

    return render(request, 'inicio_sin_sesion.html', {
        'libros_destacados': libros_destacados,
    })


@login_required
def inicio(request):
    """
    Página de inicio: muestra los últimos pedidos activos del usuario
    y estadísticas reales para el panel lateral.
    """
    cliente         = ClienteRepository.get_or_create(request.user)
    pedidos_activos = PedidoRepository.filtrar_por_estado(cliente, '').exclude(
        estado__in=['cancelado', 'entregado']
    )[:3]
    contadores      = PedidoRepository.contadores(cliente)
    ultimo_pedido   = PedidoRepository.get_all_by_cliente(cliente).first()
    num_pedidos     = contadores['todos']

    return render(request, 'inicio.html', {
        'usuario':         request.user,
        'pedidos_activos': pedidos_activos,
        'contadores':      contadores,
        'ultimo_pedido':   ultimo_pedido,
        'num_pedidos':     num_pedidos,
    })


@login_required
def perfil(request):
    cliente    = ClienteRepository.get_or_create(request.user)
    contadores = PedidoRepository.contadores(cliente)
    return render(request, 'perfil.html', {
        'usuario':    request.user,
        'contadores': contadores,
    })


def gestor(request):
    return render(request, 'gestor.html')


# ─────────────────────────────────────────
# AUTENTICATIÓN
# ─────────────────────────────────────────

def inicio_sesion(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            user = authenticate(
                request,
                username=form.cleaned_data['email'],
                password=form.cleaned_data['password'],
            )
            if user:
                login(request, user)
                return redirect('inicio')
            else:
                messages.error(request, 'Correo o contraseña incorrectos.')
    else:
        form = LoginForm()
    return render(request, 'inicio_sesion.html', {'form': form})


def registro(request):
    if request.method == 'POST':
        form = RegistroForm(request.POST)
        if form.is_valid():
            try:
                UsuarioRepository.crear(
                    email=form.cleaned_data['email'],
                    password=form.cleaned_data['password'],
                    full_name=form.cleaned_data.get('full_name', ''),
                )
                messages.success(request, '¡Cuenta creada! Ya puedes iniciar sesión.')
                return redirect('inicio_sesion')
            except ValueError as e:
                messages.error(request, str(e))
    else:
        form = RegistroForm()
    return render(request, 'registro.html', {'form': form})


# ─────────────────────────────────────────
# CATÁLOGO — API
# ─────────────────────────────────────────

def catalogo(request):
    query        = request.GET.get('q', '').strip()
    genre        = request.GET.get('genre', '')
    genre_activo = genre

    busqueda = query if query else 'bestseller'
    if genre and not query:
        busqueda = {
            'fiction': 'ficcion novela',
            'history': 'historia',
            'classic': 'clasicos literatura',
            'science': 'ciencia',
            'poetry':  'poesia',
        }.get(genre, 'literatura')

    url     = f"https://openlibrary.org/search.json?q={busqueda}&limit=30"
    headers = {'User-Agent': 'Mozilla/5.0'}
    libros  = []

    try:
        response = requests.get(url, timeout=8, headers=headers)
        if response.status_code == 200:
            data = response.json()
            for doc in data.get('docs', [])[:30]:
                titulo   = doc.get('title', 'Sin título')
                autores  = doc.get('author_name', ['Desconocido'])
                cover_id = doc.get('cover_i')
                imagen   = f"https://covers.openlibrary.org/b/id/{cover_id}-M.jpg" if cover_id else ''
                libros.append({
                    'titulo': titulo,
                    'autor':  autores[0] if autores else 'Desconocido',
                    'imagen': imagen,
                })
    except Exception as e:
        print(f'ERROR API: {e}')

    if not libros:
        libros = [
            {'titulo': 'Don Quijote de la Mancha', 'autor': 'Miguel de Cervantes',    'imagen': 'https://covers.openlibrary.org/b/id/8739161-M.jpg'},
            {'titulo': 'Cien Años de Soledad',     'autor': 'Gabriel García Márquez', 'imagen': 'https://covers.openlibrary.org/b/id/8224600-M.jpg'},
            {'titulo': 'La Sombra del Viento',     'autor': 'Carlos Ruiz Zafón',      'imagen': 'https://covers.openlibrary.org/b/id/7887754-M.jpg'},
        ]

    return render(request, 'catalogo_libros.html', {
        'libros':       libros,
        'query':        query,
        'genre_activo': genre_activo,
    })


# ─────────────────────────────────────────
# CREAR PEDIDO desde catálogo
# ─────────────────────────────────────────

@login_required
@require_POST
def crear_pedido(request):
    titulo = request.POST.get('titulo', '').strip()
    autor  = request.POST.get('autor',  '').strip()
    imagen = request.POST.get('imagen', '').strip()

    if not titulo:
        messages.error(request, 'No se pudo añadir el pedido.')
        return redirect('catalogo')

    cliente = ClienteRepository.get_or_create(request.user)
    PedidoRepository.crear(cliente=cliente, titulo=titulo, autor=autor, imagen=imagen)
    messages.success(request, f'"{titulo}" añadido a tus pedidos correctamente.')
    return redirect('mis_pedidos')


# ─────────────────────────────────────────
# MIS PEDIDOS
# ─────────────────────────────────────────

@login_required
def mis_pedidos(request):
    estado_filtro = request.GET.get('estado', '')
    cliente       = ClienteRepository.get_or_create(request.user)
    pedidos       = PedidoRepository.filtrar_por_estado(cliente, estado_filtro)
    contadores    = PedidoRepository.contadores(cliente)

    return render(request, 'mis_pedidos.html', {
        'pedidos':       pedidos,
        'estado_filtro': estado_filtro,
        'contadores':    contadores,
        'estados':       Pedido.ESTADOS,
    })


@login_required
@require_POST
def cambiar_estado_pedido(request, pedido_id):
    es_ajax      = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    cliente      = ClienteRepository.get_or_create(request.user)
    pedido       = PedidoRepository.get_by_id(pedido_id, cliente)

    if not pedido:
        return JsonResponse({'ok': False}, status=404) if es_ajax else redirect('mis_pedidos')

    nuevo_estado = request.POST.get('estado', '')
    try:
        PedidoRepository.cambiar_estado(pedido, nuevo_estado)
    except ValueError as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400) if es_ajax else redirect('mis_pedidos')

    if es_ajax:
        return JsonResponse({'ok': True, 'estado': nuevo_estado, 'label': pedido.get_estado_display()})

    messages.success(request, f'Estado actualizado: {pedido.get_estado_display()}')
    return redirect('mis_pedidos')


@login_required
@require_POST
def cancelar_pedido(request, pedido_id):
    es_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    cliente = ClienteRepository.get_or_create(request.user)
    pedido  = PedidoRepository.get_by_id(pedido_id, cliente)

    if not pedido:
        return JsonResponse({'ok': False}, status=404) if es_ajax else redirect('mis_pedidos')

    try:
        PedidoRepository.cancelar(pedido)
    except ValueError as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400) if es_ajax else redirect('mis_pedidos')

    if es_ajax:
        return JsonResponse({'ok': True})

    messages.success(request, f'Pedido "{pedido.titulo}" cancelado.')
    return redirect('mis_pedidos')


# ─────────────────────────────────────────
# EXPORTAR EXCEL
# ─────────────────────────────────────────

@login_required
def exportar_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return HttpResponse('Instala openpyxl: pip install openpyxl', status=500)

    cliente = ClienteRepository.get_or_create(request.user)
    pedidos = PedidoRepository.get_all_by_cliente(cliente)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Mis Pedidos — Librízate'

    header_font = Font(bold=True, color='FFFFFF', size=11, name='Calibri')
    header_fill = PatternFill('solid', fgColor='5E3D22')
    center      = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin        = Side(style='thin', color='D4C3B9')
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells('A1:F1')
    ws['A1'].value     = 'Librízate — Historial de Pedidos'
    ws['A1'].font      = Font(bold=True, size=14, color='5E3D22', name='Calibri')
    ws['A1'].fill      = PatternFill('solid', fgColor='FFF8F0')
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 28

    import datetime
    ws.merge_cells('A2:F2')
    ws['A2'].value     = f'Usuario: {request.user.get_full_name() or request.user.username}   |   Generado: {datetime.datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].font      = Font(size=9, color='82746C', name='Calibri')
    ws['A2'].alignment = center
    ws.row_dimensions[2].height = 16

    ws.append([])

    headers = ['#', 'Título', 'Autor', 'Estado', 'Fecha', 'Email cliente']
    for col, h in enumerate(headers, 1):
        cell           = ws.cell(row=4, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border
    ws.row_dimensions[4].height = 20

    estado_labels = dict(Pedido.ESTADOS)
    fill_par      = PatternFill('solid', fgColor='FFF8F4')
    fill_impar    = PatternFill('solid', fgColor='FFFFFF')

    for i, p in enumerate(pedidos, 1):
        ws.append([p.id, p.titulo, p.autor, estado_labels.get(p.estado, p.estado), p.fecha.strftime('%d/%m/%Y'), p.cliente.usuario.email])
        row_num = ws.max_row
        fill = fill_par if i % 2 == 0 else fill_impar
        for col in range(1, 7):
            cell           = ws.cell(row=row_num, column=col)
            cell.fill      = fill
            cell.border    = border
            cell.alignment = Alignment(vertical='center')
        ws.row_dimensions[row_num].height = 16

    ws.column_dimensions['A'].width = 7
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 13
    ws.column_dimensions['F'].width = 32

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="librizate_pedidos.xlsx"'
    return response


# ─────────────────────────────────────────
# EXPORTAR PDF
# ─────────────────────────────────────────

@login_required
def exportar_pdf(request, pedido_id):
    cliente = ClienteRepository.get_or_create(request.user)
    pedido  = PedidoRepository.get_by_id(pedido_id, cliente)

    if not pedido:
        messages.error(request, 'Pedido no encontrado.')
        return redirect('mis_pedidos')

    try:
        from weasyprint import HTML
        html_content = render(request, 'pedido_pdf.html', {'pedido': pedido}).content
        pdf = HTML(string=html_content.decode()).write_pdf()
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="pedido_{pedido.id}.pdf"'
        return response
    except ImportError:
        return render(request, 'pedido_pdf.html', {'pedido': pedido, 'print_mode': True})


# ─────────────────────────────────────────
# DETALLES
# ─────────────────────────────────────────

def add_carrito(request):
    carrito = request.session.get('carrito', [])
    carrito.append({'titulo': request.POST.get('titulo'), 'autor': request.POST.get('autor')})
    request.session['carrito'] = carrito
    return redirect('catalogo')

def detalle_pedido(request):
    return redirect('mis_pedidos')


# ─────────────────────────────────────────
# CERRAR SESIÓN
# ─────────────────────────────────────────

def cerrar_sesion(request):
    logout(request)
    return redirect('inicio_sin_sesion')