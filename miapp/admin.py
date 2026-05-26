from django.contrib import admin
from django.utils.html import format_html
from django.http import HttpResponse
from .models import Pedido, Cliente, Gestor
import io, datetime

admin.site.site_header  = "Librízate — Panel de Administración"
admin.site.site_title   = "Librízate Admin"
admin.site.index_title  = "Gestión de la Librería"


# ── CLIENTE ──────────────────────────────

@admin.register(Cliente)
class ClienteAdmin(admin.ModelAdmin):
    list_display  = ('usuario', 'telefono', 'num_pedidos')
    search_fields = ('usuario__username', 'usuario__email', 'telefono')

    @admin.display(description='Pedidos')
    def num_pedidos(self, obj):
        n = obj.pedidos.count()
        return format_html(
            '<span style="background:#ffdcc4;color:#613f24;padding:2px 10px;'
            'border-radius:999px;font-size:11px;font-weight:700">{}</span>', n
        )


# ── GESTOR ───────────────────────────────

@admin.register(Gestor)
class GestorAdmin(admin.ModelAdmin):
    list_display  = ('usuario', 'departamento', 'telefono', 'fecha_creacion')
    search_fields = ('usuario__username', 'departamento')
    list_filter   = ('departamento',)


# ── ACCIONES DE PEDIDO ───────────────────

def marcar_enviado(modeladmin, request, queryset):
    queryset.update(estado='enviado')
marcar_enviado.short_description = "✈ Marcar como Enviado"

def marcar_entregado(modeladmin, request, queryset):
    queryset.update(estado='entregado')
marcar_entregado.short_description = "✅ Marcar como Entregado"

def marcar_cancelado(modeladmin, request, queryset):
    queryset.update(estado='cancelado')
marcar_cancelado.short_description = "❌ Cancelar pedidos seleccionados"

def exportar_excel_admin(modeladmin, request, queryset):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        modeladmin.message_user(request, "Instala openpyxl: pip install openpyxl", level='error')
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Pedidos'

    headers    = ['ID', 'Cliente', 'Email', 'Título', 'Autor', 'Estado', 'Fecha']
    hdr_fill   = PatternFill('solid', fgColor='5E3D22')
    hdr_font   = Font(bold=True, color='FFFFFF', name='Calibri')
    for col, h in enumerate(headers, 1):
        c       = ws.cell(1, col, h)
        c.fill  = hdr_fill
        c.font  = hdr_font

    estado_labels = dict(Pedido.ESTADOS)
    for p in queryset.select_related('cliente__usuario'):
        ws.append([
            p.id,
            p.cliente.usuario.get_full_name() or p.cliente.usuario.username,
            p.cliente.usuario.email,
            p.titulo,
            p.autor,
            estado_labels.get(p.estado, p.estado),
            p.fecha.strftime('%d/%m/%Y %H:%M'),
        ])

    ws.column_dimensions['D'].width = 42
    ws.column_dimensions['E'].width = 28

    output = io.BytesIO()
    wb.save(output); output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="pedidos_admin.xlsx"'
    return response
exportar_excel_admin.short_description = "📊 Exportar selección a Excel"


# ── PEDIDO ───────────────────────────────

@admin.register(Pedido)
class PedidoAdmin(admin.ModelAdmin):
    list_display   = ('id', 'titulo_corto', 'autor', 'cliente_nombre', 'estado_badge', 'fecha_fmt')
    list_filter    = ('estado', 'fecha')
    search_fields  = ('titulo', 'autor', 'cliente__usuario__username', 'cliente__usuario__email')
    ordering       = ('-fecha',)
    list_per_page  = 25
    date_hierarchy = 'fecha'
    actions        = [marcar_enviado, marcar_entregado, marcar_cancelado, exportar_excel_admin]

    fieldsets = (
        ('📖 Información del libro', {
            'fields': ('titulo', 'autor', 'imagen', 'descripcion')
        }),
        ('📦 Datos del pedido', {
            'fields': ('cliente', 'estado', 'precio')
        }),
    )

    @admin.display(description='Título')
    def titulo_corto(self, obj):
        t = obj.titulo[:48] + '…' if len(obj.titulo) > 48 else obj.titulo
        return format_html('<strong>{}</strong>', t)

    @admin.display(description='Cliente')
    def cliente_nombre(self, obj):
        return obj.cliente.usuario.username

    @admin.display(description='Estado')
    def estado_badge(self, obj):
        colores = {
            'pendiente':   ('#ffdcc4', '#613f24'),
            'preparacion': ('#eedcd2', '#695c54'),
            'enviado':     ('#c0eaf2', '#244d53'),
            'entregado':   ('#d2eadd', '#0d4a25'),
            'cancelado':   ('#ffdad6', '#93000a'),
        }
        bg, fg = colores.get(obj.estado, ('#eee', '#333'))
        return format_html(
            '<span style="background:{};color:{};padding:3px 12px;border-radius:999px;'
            'font-size:11px;font-weight:700;text-transform:uppercase">{}</span>',
            bg, fg, obj.get_estado_display()
        )

    @admin.display(description='Fecha')
    def fecha_fmt(self, obj):
        return obj.fecha.strftime('%d/%m/%Y')
